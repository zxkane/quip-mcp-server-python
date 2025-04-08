import os
import csv
import io
import logging
import requests
from bs4 import BeautifulSoup
from typing import Optional, List, Dict, Any, Union
from openpyxl import load_workbook

# Initialize logger
logger = logging.getLogger("quip-mcp-server")

class QuipClient:
    """
    Simple Quip API client implementation for the MCP server
    """
    def __init__(self, access_token: str, base_url: str = "https://platform.quip.com"):
        """
        Initialize the Quip client with the given access token and base URL
        
        Args:
            access_token: Quip API access token
            base_url: Base URL for the Quip API (default: https://platform.quip.com)
        """
        self.access_token = access_token
        self.base_url = base_url.rstrip('/')
        self.session = requests.Session()
        self.session.headers = {
            'Authorization': 'Bearer ' + access_token
        }
        logger.info(f"QuipClient initialized with base URL: {self.base_url}")

    def get_thread(self, thread_id: str) -> Dict[str, Any]:
        """
        Get a thread by ID
        
        Args:
            thread_id: ID of the thread to retrieve
            
        Returns:
            Dict containing thread information
            
        Raises:
            requests.exceptions.HTTPError: If the request fails
        """
        logger.info(f"Getting thread: {thread_id}")
        response = self.session.get(f"{self.base_url}/1/threads/{thread_id}")
        response.raise_for_status()
        return response.json()

    def export_thread_to_xlsx(self, thread_id: str, output_path: str) -> str:
        """
        Export a thread to XLSX format and save it locally.
        
        Args:
            thread_id: ID of the thread to export
            output_path: Local file path where the XLSX file should be saved
            
        Returns:
            str: Path to the saved XLSX file
            
        Raises:
            requests.exceptions.HTTPError: If the request fails
        """
        logger.info(f"Exporting thread {thread_id} to XLSX")
        response = self.session.get(
            f"{self.base_url}/1/threads/{thread_id}/export/xlsx",
            stream=True  # Stream the response to handle large files
        )
        response.raise_for_status()
        
        # Ensure the directory exists
        os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
        
        # Write the file in chunks to handle large files efficiently
        with open(output_path, 'wb') as f:
            for chunk in response.iter_content(chunk_size=8192):
                if chunk:
                    f.write(chunk)
        
        logger.info(f"Successfully exported XLSX to {output_path}")
        return output_path

    def export_thread_to_csv_fallback(self, thread_id: str, sheet_name: Optional[str] = None) -> str:
        """
        Export a thread to CSV format using HTML parsing as fallback method.
        
        Args:
            thread_id: ID of the thread to export
            sheet_name: Name of the sheet to extract (optional)
            
        Returns:
            str: CSV data as string
            
        Raises:
            ValueError: If the thread is not found or does not contain a spreadsheet
        """
        logger.info(f"Using fallback method to export thread {thread_id} to CSV")
        
        # Get thread data
        thread = self.get_thread(thread_id)
        if not thread or 'html' not in thread:
            raise ValueError("Could not retrieve thread or thread has no HTML content")

        # Find and extract sheet data
        sheet = find_sheet_by_name(thread['html'], sheet_name)
        if not sheet:
            if sheet_name:
                raise ValueError(f"Could not find sheet '{sheet_name}' in the document")
            else:
                raise ValueError("Could not find any spreadsheet in the document")

        # Extract and process data
        data = extract_sheet_data(sheet)
        if not data:
            raise ValueError(f"No data found in sheet '{sheet_name or 'default'}'")

        # Convert to CSV string
        csv_buffer = io.StringIO()
        writer = csv.writer(csv_buffer, quoting=csv.QUOTE_MINIMAL)
        
        # Write all rows
        for row in data:
            writer.writerow(row)
        
        csv_data = csv_buffer.getvalue()
        csv_buffer.close()
        
        return csv_data

    def is_spreadsheet(self, thread_id: str) -> bool:
        """
        Check if a thread is a spreadsheet
        
        Args:
            thread_id: ID of the thread to check
            
        Returns:
            bool: True if the thread is a spreadsheet, False otherwise
        """
        try:
            thread = self.get_thread(thread_id)
            if not thread or 'thread' not in thread:
                return False
                
            # Check if the thread type is 'spreadsheet'
            thread_type = thread.get('thread', {}).get('type', '').lower()
            return thread_type == 'spreadsheet'
        except Exception as e:
            logger.error(f"Error checking if thread is spreadsheet: {str(e)}")
            return False


def find_sheet_by_name(document_html: str, sheet_name: Optional[str] = None) -> Optional[Any]:
    """
    Find a spreadsheet with the given name in the document HTML
    
    Args:
        document_html: HTML content of the document
        sheet_name: Name of the sheet to find (optional)
        
    Returns:
        BeautifulSoup table element or None if not found
    """
    soup = BeautifulSoup(document_html, 'html.parser')
    
    # First try to find a table with the specified title attribute
    table = soup.find('table', attrs={'title': sheet_name}) if sheet_name else None
    if table:
        return table
    
    # If not found and sheet_name is provided, look for a heading with the sheet name
    if sheet_name:
        for heading in soup.find_all(['h1', 'h2', 'h3']):
            if heading.get_text().strip() == sheet_name:
                next_table = heading.find_next('table')
                if next_table:
                    return next_table
    
    # If still not found or no sheet_name provided, return the first table
    return soup.find('table')


def is_metadata_row(row: List[str]) -> bool:
    """
    Determine if a row is likely metadata rather than a header or data row.
    
    Args:
        row: List of cell values in the row
        
    Returns:
        bool: True if the row is likely metadata, False otherwise
    """
    # Check if most cells are empty
    non_empty_cells = sum(1 for cell in row if cell.strip())
    if non_empty_cells <= 1:
        return True
        
    # Check for date patterns that suggest metadata
    date_indicators = ["updated on", "created on", "modified on", "as of"]
    row_text = " ".join(row).lower()
    if any(indicator in row_text for indicator in date_indicators):
        return True
        
    return False


def is_header_row(row: List[str]) -> bool:
    """
    Determine if a row is likely a header row.
    
    Args:
        row: List of cell values in the row
        
    Returns:
        bool: True if the row is likely a header row, False otherwise
    """
    if not row or not any(cell.strip() for cell in row):
        return False
        
    # Headers typically have:
    # 1. No very long text fields
    # 2. No common sentence punctuation
    # 3. No numbered lists or bullet points
    max_header_length = 50
    sentence_punctuation = ['.', '!', '?']
    list_indicators = ['•', '-', '1)', 'a)', '1.', 'i.', 'i)']
    
    for cell in row:
        cell = cell.strip()
        if not cell:
            continue
            
        # Check length
        if len(cell) > max_header_length:
            return False
            
        # Check for sentence punctuation (excluding abbreviations)
        if any(p in cell[1:-1] for p in sentence_punctuation):
            return False
            
        # Check for list indicators
        if any(cell.lower().startswith(indicator) for indicator in list_indicators):
            return False
    
    return True


def extract_sheet_data(sheet: Any) -> List[List[str]]:
    """
    Extract data from a sheet element
    
    Args:
        sheet: BeautifulSoup table element
        
    Returns:
        List of rows, where each row is a list of cell values
    """
    if sheet is None:
        return []
    
    # Process rows to find header and data rows
    rows_data = []
    for tr in sheet.find_all('tr'):
        cols = tr.find_all('td')
        if not cols:
            continue
            
        # Get row text
        row_text = [col.get_text().strip() for col in cols]
        
        # Skip empty rows
        if not any(text for text in row_text):
            continue
            
        # Clean up row - remove row numbers and empty columns
        cleaned_row = []
        for text in row_text:
            if text and not text.isdigit():
                # Remove any zero-width spaces and extra whitespace
                text = text.replace('\u200b', '').strip()
                cleaned_row.append(text)
        
        if cleaned_row:
            rows_data.append(cleaned_row)
    
    # Process collected rows
    if not rows_data:
        return []
        
    # Skip metadata rows at the start
    start_idx = 0
    while start_idx < len(rows_data) and is_metadata_row(rows_data[start_idx]):
        start_idx += 1
        
    if start_idx >= len(rows_data):
        return []
        
    # Find header row
    header_idx = start_idx
    while header_idx < len(rows_data):
        if is_header_row(rows_data[header_idx]):
            header_row = rows_data[header_idx]
            feature_col_idx = header_row.index('Feature to Address') if 'Feature to Address' in header_row else None
            break
        header_idx += 1
    else:
        # If no header row found, use the first non-metadata row
        header_idx = start_idx - 1
        feature_col_idx = None
    
    # Create the final rows list
    rows = []
    
    # Include any metadata rows at the start
    for i in range(start_idx):
        rows.append(rows_data[i])
    
    # Add header row if found
    if header_idx >= start_idx:
        rows.append(rows_data[header_idx])
    
    # Add data rows (process special formatting for Feature to Address column)
    for row_data in rows_data[header_idx + 1:]:
        if not is_metadata_row(row_data):
            processed_row = []
            for col_idx, cell in enumerate(row_data):
                cell_text = cell.strip()
                
                # Special handling for Feature to Address column
                if feature_col_idx is not None and col_idx == feature_col_idx:
                    # Handle lists specially
                    if cell_text.startswith('a)'):
                        # Already in a)b)c) format, just ensure proper newlines
                        for i in range(97, 122):  # ASCII 'a' to 'z'
                            char = chr(i)
                            next_char = chr(i+1)
                            cell_text = cell_text.replace(f"{char}){next_char})", f"{char})\n{next_char})")
                
                processed_row.append(cell_text)
            
            if any(cell.strip() for cell in processed_row):  # Only add non-empty rows
                rows.append(processed_row)
    
    return rows


def convert_xlsx_to_csv(xlsx_path: str, sheet_name: Optional[str] = None) -> str:
    """
    Convert XLSX file to CSV format, optionally extracting a specific sheet.
    
    Args:
        xlsx_path: Path to the XLSX file
        sheet_name: Name of the sheet to extract (optional)
        
    Returns:
        str: CSV data as string
        
    Raises:
        ValueError: If the sheet is not found
    """
    logger.info(f"Reading XLSX file from {xlsx_path}")
    
    # Load the workbook with all data
    wb = load_workbook(filename=xlsx_path, read_only=False, data_only=True)
    
    # Get available sheet names
    sheet_names = wb.sheetnames
    logger.info(f"Available sheets: {', '.join(sheet_names)}")
    
    # Determine which sheet to use
    target_sheet = None
    if sheet_name:
        # Try exact match first
        if sheet_name in sheet_names:
            target_sheet = wb[sheet_name]
        else:
            # Try case-insensitive match
            sheet_lower = sheet_name.lower()
            for s in sheet_names:
                if s.lower() == sheet_lower:
                    target_sheet = wb[s]
                    break
            
            if not target_sheet:
                raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {', '.join(sheet_names)}")
    else:
        # Use first sheet if no name specified
        target_sheet = wb.active
    
    # Convert to CSV
    csv_buffer = io.StringIO()
    csv_writer = csv.writer(csv_buffer)
    
    # Get all column letters from sheet dimensions
    from openpyxl.utils import get_column_letter
    max_col = 0
    for col in target_sheet.columns:
        max_col = max(max_col, col[0].column)
    
    logger.info(f"Found {max_col} columns")
    column_letters = [get_column_letter(i) for i in range(1, max_col + 1)]
    
    # Process each row
    for row_idx, row in enumerate(target_sheet.rows, 1):
        # Get values for all possible columns
        row_data = []
        for col in column_letters:
            cell = target_sheet[f"{col}{row_idx}"]
            value = cell.value
            if isinstance(value, str):
                value = value.strip()
            row_data.append('' if value is None else str(value))
            
        # Log first few rows for debugging
        if row_idx <= 5:
            logger.info(f"Row {row_idx} data: {row_data}")
            
        csv_writer.writerow(row_data)
    
    # Get the CSV data
    csv_data = csv_buffer.getvalue()
    csv_buffer.close()
    wb.close()
    
    return csv_data
